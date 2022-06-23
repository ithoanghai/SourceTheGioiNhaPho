import csv
import math
import os
import tempfile
from datetime import date, datetime

import xlsxwriter
import requests
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_protect
from openpyxl import load_workbook, Workbook
from rest_framework import request, response
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.authentication import SessionAuthentication, BasicAuthentication

from FunctionModule.listings.export_csv import prepare_fb_headers, prepare_fb_listing_data
from FunctionModule.listings.models import ListingSerializer, Listing
from FunctionModule.listings.search import prepare_listing_queryset, search_by_keywords, get_suggestions
from FunctionModule.realtors.export_realtor import prepare_export_realtor_data
from FunctionModule.realtors.models import Realtor


@api_view(['GET'])
@authentication_classes([])
@permission_classes([])
def get_states(r: request.Request, **kwargs):
    query = kwargs.get('lookup', '')
    action = kwargs.get('action', 'sell')
    data = get_suggestions(query)
    return response.Response(data)


@api_view(['GET'])
@authentication_classes([SessionAuthentication, BasicAuthentication])
def get_user(r: request.Request, **kwargs):

    return response.Response({
        "username": r.user.username,
        "can_export_listing": r.user.is_superuser,
        "can_import_listing": r.user.is_superuser,
        "can_export_realtor": r.user.is_superuser,
        "can_import_realtor": r.user.is_superuser,
    })


def query_params_to_filters(input_params):
    query = []

    is_verified = input_params.get('is_verified', None)
    if is_verified:
        query.append(['is_verified=true'])

    is_exclusive = input_params.get('is_exclusive', None)
    if is_exclusive:
        query.append(['is_exclusive=true'])

    # Status: List
    status: str = input_params.get('status', None)
    if status:
        statuses = status.split(',')
        if len(statuses):
            st_query = [f'status={s}' for s in statuses]
            query.append(st_query)

    # Directions: List
    directions = input_params.get('directions', None)
    if directions:
        directions = directions.split(',')
        if len(directions):
            direction_query = [f'status={s}' for s in directions]
            query.append(direction_query)

    # House type: List
    house_type = input_params.get('house_type', None)
    if house_type:
        house_type = house_type.split(',')
        if len(house_type):
            ht_query = [f'status={s}' for s in house_type]
            query.append(ht_query)

    # Bedrooms: List
    bedrooms = input_params.get('bedrooms', None)
    if bedrooms:
        bedrooms = bedrooms.split(',')
        to_filter = []
        for f in bedrooms:
            if f == '6+':
                to_filter.append('bedrooms>=6')
            else:
                to_filter.append(f'bedrooms={f}')
        if len(to_filter):
            query.append(to_filter)

    # Bathrooms: Single string
    bathrooms = input_params.get('bathrooms', None)
    if bathrooms:
        if bathrooms == 'all':
            pass
        elif bathrooms == '5+':
            query.append(['bathrooms>=5'])
        else:
            query.append([f'bathrooms={bathrooms}'])

    # Price: minPrice and maxPrice each single number
    price = input_params.get('minPrice', None)
    if price and price.isnumeric():
        query.append([f'price>={price}'])
    max_price = input_params.get('maxPrice', None)
    if max_price and max_price.isnumeric():
        query.append([f'price<{max_price}'])

    # Area: minArea and maxArea each single number
    area = input_params.get('minArea')
    if area and area.isnumeric():
        query.append([f'area>={area}'])
    max_area = input_params.get('maxArea')
    if max_area and max_area.isnumeric():
        query.append([f'area>={max_area}'])

    if 'sort' in input_params:
        sort_by = input_params.get('sort')

    return query


@api_view(['GET'])
@authentication_classes([])
@permission_classes([])
def search_listing(req: request.Request, **kwargs):
    try:
        page = int(req.query_params.get('page', 1))
    except ValueError:
        page = 1

    previous_page = None if page == 1 else page - 1
    limit = req.query_params.get('limit', 50)
    offset = (page - 1) * limit
    min_page = page - 1 if page - 1 > 0 else page

    # keywords = req.query_params.get('keywords', '')
    # if keywords:
    #     filters = query_params_to_filters(req.query_params)
    #     results = search_by_keywords(keywords, limit, offset, filters)
    #     total: int = results['nbHits']
    #     total_pages = int(math.ceil(total / limit))
    #     next_page = page + 1 if total_pages > page else None
    #     next_5_pages = list(range(min_page, page + 5 if page + 5 < total_pages else total_pages))
    #     next_5_pages = next_5_pages if len(next_5_pages) > 1 else [page]
    #
    #     return response.Response({
    #         'listings': results['hits'],
    #         'pagination': {
    #             'current_page': page,
    #             'previous_page': previous_page,
    #             'next_page': next_page,
    #             'next_5_pages': next_5_pages,
    #             'limit': limit,
    #             'offset': offset,
    #             'total': results['nbHits']
    #         }
    #     })

    queryset_list = prepare_listing_queryset(req.query_params)
    paginator = Paginator(queryset_list, limit)
    total_pages = paginator.num_pages
    next_page = page + 1 if total_pages > page else None
    next_5_pages = list(range(min_page, page + 5 if page + 5 < total_pages else total_pages))
    next_5_pages = next_5_pages if len(next_5_pages) > 1 else [page]

    try:
        listings = paginator.get_page(page)
    except Exception:
        listings = paginator.get_page(1)

    return response.Response({
        "listings": ListingSerializer(listings, many=True).data,
        # "listings": [],
        "pagination": {
            'current_page': page,
            'previous_page': previous_page,
            'next_page': next_page,
            'next_5_pages': next_5_pages,
            'limit': limit,
            'offset': offset,
            'total': queryset_list.count()
        }
    })


@api_view(['POST'])
@csrf_protect
def download_exported_listing(req: request.Request, **kwargs):
    file_path = os.path.join(tempfile.gettempdir(), 'tmp.csv')
    with open(file_path, 'w', encoding='utf-8') as fp:
        headers = prepare_fb_headers()
        writer = csv.DictWriter(fp, fieldnames=headers)
        writer.writeheader()
        for listing in Listing.objects.all():
            listing_data = prepare_fb_listing_data(listing)
            writer.writerow(listing_data)
    with open(file_path, 'r', encoding='utf-8') as fp:
        resp = HttpResponse(fp.read(), content_type="text/csv")
        resp['Content-Disposition'] = f'filename=tgnp_bds_facebook_export-{datetime.today().strftime("%Y-%m-%d")}.csv'
        print("Exported listings")
        return resp


@api_view(['POST'])
@csrf_protect
def download_export_realtor(req: request.Request, **kwargs):
    #file_path = os.path.join(tempfile.gettempdir(), 'tmp.xlsx')
    file_path = os.path.join('./media/import-export/', 'DS NHÂN SỰ TGNP.xlsx')
    #wb = Workbook()
    #ws = wb.create_sheet("DS Chuyên viên TGNP")
    #ws.title = "DS Chuyên viên TGNP"
    #wb.save(file_path)
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet('DS TOÀN BỘ CHUYÊN VIÊN TGNP')

    # write header
    row = 0
    worksheet.write(row, 0, 'Họ và Tên')
    worksheet.write(row, 1, 'Ngày sinh')
    worksheet.write(row, 2, 'Số Điện Thoại')
    worksheet.write(row, 3, 'CMT/CCCD')
    worksheet.write(row, 4, 'Ngày vào đơn vị')
    worksheet.write(row, 5, 'Chức vụ')
    worksheet.write(row, 6, 'Đơn vị')
    worksheet.write(row, 7, 'Bộ phận/Phòng/Ban')
    worksheet.write(row, 8, 'Facebook')
    worksheet.write(row, 9, 'Email')
    worksheet.write(row, 10, 'Quê quán')
    worksheet.write(row, 11, 'Nơi ở')
    worksheet.write(row, 12, 'Địa bàn hoạt động')
    worksheet.write(row, 13, 'Hiện trạng')
    worksheet.write(row, 14, 'Hợp tác với TGNP')
    worksheet.write(row, 15, 'Công khai danh tính')
    worksheet.write(row, 16, 'Khóa đào tạo')
    worksheet.write(row, 17, 'Nguồn tuyển/Giới thiệu')

    row = 1
    column = 0
    realtors = Realtor.objects.all()
    # iterating through content list
    for realtor in realtors:
        # write operation perform
        worksheet.write(row, 0, realtor.name)
        format1 = workbook.add_format({'num_format': 'd-m-yyyy'})
        worksheet.write(row, 1, realtor.birthyear, format1)
        worksheet.write(row, 2, realtor.phone1)
        worksheet.write(row, 3, realtor.identifier)
        #worksheet.write_datetime(row, 4, realtor.date_join)
        worksheet.write(row, 5, realtor.position)
        worksheet.write(row, 6, realtor.workplace)
        worksheet.write(row, 7, realtor.department)
        worksheet.write(row, 8, realtor.facebook)
        worksheet.write(row, 9, realtor.email)
        worksheet.write(row, 10, realtor.countryside)
        worksheet.write(row, 11, realtor.address)
        worksheet.write(row, 12, realtor.work_area)
        worksheet.write(row, 13, realtor.status)
        worksheet.write(row, 14, realtor.is_cooperate)
        worksheet.write(row, 15, realtor.is_published)
        worksheet.write(row, 16, realtor.training)
        worksheet.write(row, 17, realtor.referral)

        # incrementing the value of row by one
        # with each iterations.
        row += 1

    #download(file_path, dest_folder="C:/")
    workbook.close()
    return response.Response()


def download(url: str, dest_folder: str):
    if not os.path.exists(dest_folder):
        os.makedirs(dest_folder)  # create folder if it does not exist

    filename = url.split('/')[-1].replace(" ", "_")  # be careful with file names
    file_path = os.path.join(dest_folder, filename)

    r = requests.get(url, stream=True)
    if r.ok:
        print("saving to", os.path.abspath(file_path))
        with open(file_path, 'wb') as f:
            for chunk in r.iter_content(chunk_size=1024 * 8):
                if chunk:
                    f.write(chunk)
                    f.flush()
                    os.fsync(f.fileno())
    else:  # HTTP status code 4XX/5XX
        print("Download failed: status code {}\n{}".format(r.status_code, r.text))