import io
import os
import warnings
from datetime import datetime

import xlsxwriter
from django.utils.translation import gettext_lazy as _
from django.contrib import admin
from django.core.exceptions import PermissionDenied
from django.db.models import Q
from django.template.response import TemplateResponse
from import_export import resources
from import_export.admin import ImportExportModelAdmin
from import_export.forms import ImportExportFormBase
from import_export.signals import post_export
from openpyxl import load_workbook

from .choices import Position, Title, Workplace
from .forms import ImportForm, ExportForm
from .models import Realtor
from django.http import HttpRequest, JsonResponse, HttpResponse

from TownhouseWorldRealestate.filters import ChoicesFieldListFilter, DateFieldListFilter, BooleanFieldListFilter  # , RangeNumericFilter
from ..customers.models import Customer
from ..listings import TransactionType, print_trace, Status
from ..listings.models import ListingHistory, Listing
from ..transactions.models import Transaction, TransactionHistory


def read_header(header_row):
    header_dict = {
        "tieude": 0,
        "mota": 1,
        "mota-dauchu": 2,
        "anh-nha": 3,
        "anh-so": 4,
        "tgian": 5,
        "hien-trang": 6,
        "dia-chi": 7,
        "pho": 8,
        "quan": 9,
        "thong-so": 10,
        "gia": 11,
        "dau-chu": 12,
        "sdt": 13,
        "don-vi": 14,
        "dac-diem": 15,
        "huong": 16,
        "dt": 20,
        "trm2": 21,
        "hientrang": 22,
        # "thanh-pho": 20,
        # "so-tang": 23,
        # "mat-tien": 23,
        # "hoa-hong": 23,
        # "nguon": 23,
    }

    return header_dict


class RealtorResource(resources.ModelResource):

    class Meta:
        model = Realtor


class ListingInline(admin.TabularInline):
    model = Listing
    extra = 0  # If you have a fixed number number of answers, set it here.
    fields = ('date_created','date_update', 'status', 'address', 'area','floors', 'price','is_published' )
    show_change_link = True
    verbose_name = "BẤT ĐỘNG SẢN CẦN QUẢN LÝ"
    verbose_name_plural = "20 BĐS CŨ NHẤT CỦA CHUYÊN VIÊN CÒN GIAO DỊCH"

    def get_queryset(self, request):
        """Alter the queryset to return no existing entries"""
        # get the existing query set, then empty it.
        LIMIT_SEARCH = 20
        queryset = super(ListingInline, self).get_queryset(request)
        ids = queryset.order_by('date_created').values('pk')[:LIMIT_SEARCH]
        qs = Listing.objects.filter(~Q(status=Status.SOLD),pk__in=ids).order_by('date_created')
        #return qs.filter(user=request.user)
        return qs


class CustomerInline(admin.TabularInline):
    model = Customer
    extra = 0  # If you have a fixed number number of answers, set it here.
    fields = ('hire_date','financial_range',  'custormer_type', 'transactionStatus', 'name', 'phone', 'district', )
    show_change_link = True
    verbose_name = "KHÁCH HÀNG CẦN QUẢN LÝ"
    verbose_name_plural = "DS KHÁCH HÀNG CẦN CHĂM SÓC"

    def get_queryset(self, request):
        """Alter the queryset to return no existing entries"""
        # get the existing query set, then empty it.
        LIMIT_SEARCH = 20
        queryset = super(CustomerInline, self).get_queryset(request)
        ids = queryset.order_by('-financial_range').values('pk')[:LIMIT_SEARCH]
        qs = Customer.objects.filter(pk__in=ids).order_by('-financial_range')
        # return qs.filter(user=request.user)
        return qs


class TransactionInline(admin.TabularInline):
    model = Transaction
    extra = 0  # If you have a fixed number number of answers, set it here.
    fields = ('date', 'status', 'trantype', 'message', 'request_price', 'customer',)
    verbose_name = "GIAO DỊCH LIÊN QUAN"
    verbose_name_plural = "GIAO DỊCH ĐƯỢC GIAO QUẢN LÝ"
    readonly_fields = ('status', 'trantype', 'message', 'request_price', 'customer',)
    can_delete = False

    def get_queryset(self, request):
        """Alter the queryset to return no existing entries"""
        # get the existing query set, then empty it.
        LIMIT_SEARCH = 20
        queryset = super(TransactionInline, self).get_queryset(request)
        ids = queryset.order_by('-date').values('pk')[:LIMIT_SEARCH]
        qs = Transaction.objects.filter(pk__in=ids).order_by('-date')
        # return qs.filter(user=request.user)
        return qs


class TransactionHistoryInline(admin.TabularInline):
    model = TransactionHistory
    extra = 0  # If you have a fixed number number of answers, set it here.
    fields = ('date', 'status', 'transaction', 'reason', 'comment',  'realtor', )
    verbose_name = "XỬ LÝ GIAO DỊCH"
    verbose_name_plural = "XỬ LÝ GIAO DỊCH"
    show_change_link = True
    readonly_fields = ()
    can_delete = False

    def get_queryset(self, request):
        """Alter the queryset to return no existing entries"""
        # get the existing query set, then empty it.
        LIMIT_SEARCH = 20
        queryset = super(TransactionHistoryInline, self).get_queryset(request)
        ids = queryset.order_by('-date').values('pk')[:LIMIT_SEARCH]
        qs = TransactionHistory.objects.filter(pk__in=ids).order_by('-date')
        # return qs.filter(user=request.user)
        return qs


class RealtorAdmin(ImportExportModelAdmin):
    resource_classes = [RealtorResource]
    import_form_class = ImportForm
    export_form_class = ExportForm

    list_display = ('join_date','name', 'position', 'phone1', 'department', 'work_area','is_cooperate','status','is_published', 'facebook')
    list_display_links = ('name',)
    search_fields = ('id', 'name','phone1','phone2', 'email', 'address','title','birthyear','identifier', 'position', 'countryside', 'workplace', 'department', 'work_area', 'facebook','referral', 'hire_date')
    autocomplete_fields = ['user']
    list_per_page = 50
    ordering = ('-name', '-hire_date')
    sortable_by = ('position')
    #filter_vertical = ('user',)
    list_editable = ('is_cooperate','is_published',)
    list_filter = (
        ('position', ChoicesFieldListFilter),
        ('workplace', ChoicesFieldListFilter),
        ('is_cooperate', BooleanFieldListFilter),
        ('is_published', BooleanFieldListFilter),
        ('status', ChoicesFieldListFilter),
        ('hire_date', DateFieldListFilter),
        ('date_join', DateFieldListFilter),
        # ('birthyear', RangeNumericFilter),
        # ('level', RangeNumericFilter),
    )
    readonly_fields = []
    inlines = [CustomerInline, ListingInline, TransactionInline, TransactionHistoryInline,   ]
    actions = ['make_published', 'unpublished']

    add_fieldsets = (
        ('THÔNG TIN CƠ BẢN CHUYÊN VIÊN', {
            'classes': ('wide',),
            'fields': ('user', ( 'name', 'birthyear', 'position'), ('avatar', 'countryside', 'address'), ('phone1', 'phone2','email')),
        }), ('THÔNG TIN BỔ SUNG', {'fields': (
            ( 'workplace', 'department', 'work_area','identifier',),
            'facebook', 'story',)})
    )

    fieldsets = (
        ('THÔNG TIN CƠ BẢN CHUYÊN VIÊN', {
            'classes': ('wide',),
            'fields': ('user', ('name', 'birthyear', 'position'), ('avatar','countryside', 'address'), ('phone1', 'phone2', 'email')),
        }),
        ('THÔNG TIN BỔ SUNG', {'fields': (
            ('workplace', 'department', 'work_area','identifier',),
            ('facebook', 'website', 'youtube'),
            ('title', 'level', 'training','referral',),
            ('story'))}),
        ('THÔNG TIN KHÁC', {
            'fields': (('status', 'is_cooperate', 'is_published'),)
        }),
        ('THỜI GIAN HOẠT ĐỘNG', {'fields': (('date_join', 'hire_date'),)}),
    )

    def changelist_view(
        self,
        request: HttpRequest,
        form_url: str = '',
        extra_context: dict = None
    ) -> HttpResponse:

        extra_context = extra_context or {}
        extra_context.update({
            'extrabutton': True,
        })
        return super().changelist_view(request, extra_context=extra_context)

    def make_published(self, request, queryset):
        updated = queryset.update(is_published=True)
        self.message_user(request, f'Chuyển sang chế độ đăng công khai danh tính cho {updated}')

    def unpublished(self, request, queryset):
        updated = queryset.update(is_published=False)
        self.message_user(request, f'Đã đổi trạng thái ẩn danh tính cho {updated}')

    def import_action(self, request, *args, **kwargs):
        context = self.get_import_context_data()

        import_formats = self.get_import_formats()
        if getattr(self.get_form_kwargs, "is_original", False):
            # Use new API
            import_form = self.create_import_form(request)
        else:
            form_class = self.get_import_form_class(request)
            form_kwargs = self.get_form_kwargs(form_class, *args, **kwargs)

            if issubclass(form_class, ImportExportFormBase):
                import_form = form_class(
                    import_formats,
                    self.get_import_resource_classes(),
                    request.POST or None,
                    request.FILES or None,
                    **form_kwargs
                )
            else:
                warnings.warn(
                    "The ImportForm class must inherit from ImportExportFormBase, "
                    "this is needed for multiple resource classes to work properly. ",
                    category=DeprecationWarning
                )
                import_form = form_class(
                    import_formats,
                    request.POST or None,
                    request.FILES or None,
                    **form_kwargs
                )

        resources = list()
        if request.POST and import_form.is_valid():
            input_format = import_formats[int(import_form.cleaned_data['input_format'])]()
            if not input_format.is_binary():
                input_format.encoding = self.from_encoding
            import_file = import_form.cleaned_data['import_file']
            # first always write the uploaded file to disk as it may be a
            # memory file or else based on settings upload handlers
            tmp_storage = self.write_to_tmp_storage(import_file, input_format)
            # allows get_confirm_form_initial() to include both the
            # original and saved file names from form.cleaned_data
            import_file.tmp_storage_name = tmp_storage.name

            try:
                # then read the file, using the proper format-specific mode
                # warning, big files may exceed memory
                data = tmp_storage.read()
                dataset = input_format.create_dataset(data)
            except Exception as e:
                import_form.add_error('import_file',
                                      _(f"'{type(e).__name__}' encountered while trying to read file. "
                                        "Ensure you have chosen the correct format for the file. "
                                        f"{str(e)}"))

            if not import_form.errors:
                # prepare kwargs for import data, if needed
                res_kwargs = self.get_import_resource_kwargs(request, form=import_form, *args, **kwargs)
                resource = self.choose_import_resource_class(import_form)(**res_kwargs)
                resources = [resource]

                if request.POST.get('import_type') == 'Database':
                    # prepare additional kwargs for import_data, if needed
                    imp_kwargs = self.get_import_data_kwargs(request, form=import_form, *args, **kwargs)
                    result = resource.import_data(dataset, dry_run=True,
                                                  raise_errors=False,
                                                  file_name=import_file.name,
                                                  user=request.user,
                                                  **imp_kwargs)
                    context['result'] = result

                    if not result.has_errors() and not result.has_validation_errors():
                        if getattr(self.get_form_kwargs, "is_original", False):
                            # Use new API
                            context["confirm_form"] = self.create_confirm_form(
                                request, import_form=import_form
                            )
                        else:
                            confirm_form_class = self.get_confirm_form_class(request)
                            initial = self.get_confirm_form_initial(request, import_form)
                            context["confirm_form"] = confirm_form_class(
                                initial=self.get_form_kwargs(form=import_form, **initial)
                            )
                elif request.POST.get('import_type') == 'Custom':
                    line_count = 0
                    try:
                        if request.user is not None:
                            user = request.user
                            query = Q(phone1=user.phone) or Q(user_id=user.id)
                            real = Realtor.objects.filter(query)
                            if real.first() is None:
                                Realtor.objects.create(user_id=user.id, phone1=user.phone, name=user.name, email=user.email,
                                                       address=user.address, is_cooperate=True)
                        workbook = load_workbook(import_file)
                        # Define variable to read the active sheet:
                        worksheet = workbook.active
                        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                            line_count += 1
                            realtor = Realtor()
                            cols = [c.value for c in row]
                            if cols[0] is not None:
                                if "khối" in cols[0].lower():
                                    name = cols[0].lower().split('khối')
                                    realtor.name = name[0]
                                    realtor.department = "khối " + name[1]
                                elif "k." in cols[0].lower():
                                    name = cols[0].lower().split('k.')
                                    realtor.name = name[0]
                                    realtor.department = "khối " + name[1]
                                elif "phòng" in cols[0].lower():
                                    name = cols[0].lower().split('phòng')
                                    realtor.name = name[0]
                                    realtor.department = "phòng " + name[1]
                                elif "p." in cols[0].lower():
                                    name = cols[0].lower().split('p.')
                                    realtor.name = name[0]
                                    realtor.department = "phòng " + name[1]
                                else:
                                    realtor.name = cols[0]

                            if cols[1] is not None:
                                year = cols[1].strftime("%Y")
                                realtor.birthyear = int(str(year))
                            if cols[2] is not None:
                                phone = cols[2].split(' ')
                                realtor.phone1 = phone[0]
                                if len(phone) > 1:
                                    realtor.phone2 = phone[1]
                            else:
                                realtor.phone1 = realtor.name
                            realtor.identifier = cols[3]
                            if cols[4] is not None:
                                realtor.date_join = cols[4].strftime("%Y-%m-%d")
                            if cols[5] is not None:
                                if "trợ lý" in cols[5].lower():
                                    realtor.position = Position.ASSISTANT
                                    realtor.title = Title.COMPETENTLY
                                    realtor.level = 5
                                elif "thư ký" in cols[5].lower():
                                    realtor.position = Position.SECRETARY
                                    realtor.title = Title.COMPETENTLY
                                    realtor.level = 6
                                elif "bá tước" in cols[5].lower():
                                    realtor.position = Position.REGIONAL_DIRECTOR
                                    realtor.title = Title.EXPERT
                                    realtor.level = 7
                                elif "giám đốc" in cols[5].lower():
                                    realtor.position = Position.MANAGING_DIRECTOR
                                    realtor.title = Title.EXPERT
                                    realtor.level = 9
                                elif "trưởng phòng" in cols[5].lower() or "uvtp" in cols[5].lower():
                                    realtor.position = Position.MANAGER
                                    realtor.title = Title.EXPERT
                                    realtor.level = 8
                                elif "đầu chủ" in cols[5].lower() or "ứng viên đầu chủ" in cols[5].lower():
                                    realtor.position = Position.EXPERT_HOME
                                    realtor.title = Title.MASTER
                                    realtor.level = 7
                                else:
                                    realtor.position = cols[5].lower()
                            if cols[6] is not None:
                                if "nhà phố việt nam" in cols[6].lower():
                                    realtor.workplace = Workplace.NHAPHO
                                elif "thế giới nhà phố" in cols[6].lower():
                                    realtor.workplace = Workplace.TGNP
                                elif "thiên khôi" in cols[6].lower():
                                    realtor.workplace = Workplace.THIENKHOI
                                elif "tuấn 123" in cols[6].lower():
                                    realtor.workplace = Workplace.TUAN123
                                elif "kimland" in cols[6].lower():
                                    realtor.workplace = Workplace.KIMLAND
                                elif "sumo" in cols[6].lower():
                                    realtor.workplace = Workplace.SUMO
                                elif "vuda" in cols[6].lower():
                                    realtor.workplace = Workplace.VUDA
                                elif "khác" in cols[6].lower():
                                    realtor.workplace = Workplace.OTHER
                                else:
                                    realtor.workplace = cols[6].lower()
                            if cols[7] is not None:
                                realtor.department = cols[7]
                            realtor.facebook = cols[8]
                            realtor.email = cols[9]
                            realtor.countryside = cols[10]
                            realtor.address = cols[11]
                            realtor.work_area = cols[12]
                            if cols[13] is not None:
                                if "đang hợp tác" in cols[13].lower():
                                    realtor.status = Status.COOPERATING
                                elif "tạm dừng" in cols[13].lower():
                                    realtor.status = Status.PAUSE_COOPERATING
                                elif "ngừng hợp tác" in cols[13].lower() or "dừng hợp tác" in cols[13].lower():
                                    realtor.status = Status.STOP_COOPERATING
                                else:
                                    realtor.status = cols[13]
                            if cols[14] is not None:
                                if type(cols[14]) == bool:
                                    realtor.is_cooperate = cols[14]
                                elif "đang hợp tác" in cols[14]:
                                    realtor.is_cooperate = True
                                else:
                                    realtor.is_cooperate = False
                            if cols[15] is not None:
                                if type(cols[15]) == bool:
                                    realtor.is_published = cols[15]
                                elif "có" in cols[15].lower():
                                    realtor.is_published = True
                                else:
                                    realtor.is_cooperate = False
                            realtor.training = cols[16]
                            realtor.referral = cols[17]

                            query = Q(name=realtor.name)
                            if realtor.phone1 is not None:
                                query = query or Q(phone1=realtor.phone1)
                            if realtor.email is not None:
                                query = query or Q(email=realtor.email)
                            if realtor.facebook is not None:
                                query = query or Q(facebook=realtor.facebook)

                            if realtor.phone1 != realtor.name:
                                real = Realtor.objects.filter(phone1=realtor.phone1).first()
                                if real is None:
                                    if realtor.phone1 is not None:
                                        realtor.save()
                                        print(f"Thêm Realtor {realtor} chưa có")
                                else:
                                    if realtor.email is not None:
                                        query = query or Q(email=realtor.email)
                                    if realtor.phone1 is not None:
                                        query = query or Q(phone1=realtor.phone1) or Q(phone2=realtor.phone1) or Q(
                                            story__contains=realtor.phone1)
                                    if realtor.phone2 is not None:
                                        query = query or Q(phone2=realtor.phone2) or Q(phone1=realtor.phone2) or Q(
                                            story__contains=realtor.phone2)
                                    if Realtor.objects.filter(query).first() is not None:
                                        real = Realtor.objects.filter(query).first()
                                    real.name = realtor.name
                                    real.phone1 = realtor.phone1
                                    real.identifier = realtor.identifier
                                    real.birthyear = realtor.birthyear
                                    real.date_join = realtor.date_join
                                    real.position = realtor.position
                                    real.workplace = realtor.workplace
                                    real.department = realtor.department
                                    real.facebook = realtor.facebook
                                    real.email = realtor.email
                                    real.countryside = realtor.countryside
                                    real.work_area = realtor.work_area
                                    real.status = realtor.status
                                    real.is_cooperate = realtor.is_cooperate
                                    real.is_published = realtor.is_published
                                    real.save()
                                    print(f"Cập nhật realtor {real}")
                            elif not Realtor.objects.filter(query).exists():
                                realtor.save()
                                print(f"Thêm Realtor {realtor} chưa có số")
                            elif realtor.phone1 == realtor.name:
                                reals = Realtor.objects.filter(facebook=realtor.facebook)
                                if reals.count() > 1:
                                    for r in reals:
                                        if not r.phone1.isalnum():
                                            r.delete()

                        workbook.close()

                    except Exception as ex:
                        print(f"Error occurred at line: {line_count} type {type(ex)}")
                        print_trace(ex)
                    finally:
                        print(f'Read {line_count} lines')

                if not self.has_import_permission(request):
                    raise PermissionDenied
        else:
            res_kwargs = self.get_import_resource_kwargs(request, form=import_form, *args, **kwargs)
            resource_classes = self.get_import_resource_classes()
            resources = [resource_class(**res_kwargs) for resource_class in resource_classes]

        context.update(self.admin_site.each_context(request))

        context['title'] = _("Import")
        context['form'] = import_form
        context['opts'] = self.model._meta
        context['media'] = self.media + import_form.media
        context['fields_list'] = [
            (resource.get_display_name(), [f.column_name for f in resource.get_user_visible_fields()])
            for resource in resources
        ]

        request.current_app = self.admin_site.name
        return TemplateResponse(request, [self.import_template_name],
                                context)

    def export_action(self, request, *args, **kwargs):
        if not self.has_export_permission(request):
            raise PermissionDenied

        if getattr(self.get_export_form, 'is_original', False):
            form_type = self.get_export_form_class()
        else:
            form_type = self.get_export_form()
        formats = self.get_export_formats()
        form = form_type(formats, self.get_export_resource_classes(), request.POST or None)
        if form.is_valid():
            file_format = formats[
                int(form.cleaned_data['file_format'])
            ]()
            content_type = file_format.get_content_type()
            if request.POST.get('export_type') == 'Database':
                queryset = self.get_export_queryset(request)
                export_data = self.get_export_data(
                    file_format, queryset, request=request, encoding=self.to_encoding, export_form=form,
                )
                response = HttpResponse(export_data, content_type=content_type)
                response['Content-Disposition'] = 'attachment; filename="%s"' % (
                    self.get_export_filename(request, queryset, file_format),
                )

                post_export.send(sender=None, model=self.model)
                return response
            elif request.POST.get('export_type') == 'Custom':
                output = io.BytesIO()
                workbook = xlsxwriter.Workbook(output, {'in_memory': True})
                worksheet = workbook.add_worksheet('DS TOÀN BỘ CHUYÊN VIÊN TGNP')
                # write header
                row = 0
                worksheet.write(row, 0, 'Họ và Tên')
                worksheet.write(row, 1, 'Ngày sinh')
                worksheet.write(row, 2, 'Số Điện Thoại')
                worksheet.write(row, 3, 'CMT/CCCD')
                worksheet.write(row, 4, 'Ngày vào đơn vị')
                worksheet.write(row, 5, 'Chức vụ')
                worksheet.write(row, 6, 'Đơn vị')
                worksheet.write(row, 7, 'Bộ phận/Phòng/Ban')
                worksheet.write(row, 8, 'Facebook')
                worksheet.write(row, 9, 'Email')
                worksheet.write(row, 10, 'Quê quán')
                worksheet.write(row, 11, 'Nơi ở')
                worksheet.write(row, 12, 'Địa bàn hoạt động')
                worksheet.write(row, 13, 'Hiện trạng')
                worksheet.write(row, 14, 'Hợp tác với TGNP')
                worksheet.write(row, 15, 'Công khai danh tính')
                worksheet.write(row, 16, 'Khóa đào tạo')
                worksheet.write(row, 17, 'Nguồn tuyển/Giới thiệu')

                row = 1
                column = 0
                realtors = Realtor.objects.all()
                # iterating through content list
                for realtor in realtors:
                    # write operation perform
                    worksheet.write(row, 0, realtor.name)
                    format1 = workbook.add_format({'num_format': 'd-m-yyyy'})
                    worksheet.write(row, 1, realtor.birthyear, format1)
                    worksheet.write(row, 2, realtor.phone1)
                    worksheet.write(row, 3, realtor.identifier)
                    worksheet.write(row, 4, realtor.date_join, format1)
                    worksheet.write(row, 5, realtor.position)
                    worksheet.write(row, 6, realtor.workplace)
                    worksheet.write(row, 7, realtor.department)
                    worksheet.write(row, 8, realtor.facebook)
                    worksheet.write(row, 9, realtor.email)
                    worksheet.write(row, 10, realtor.countryside)
                    worksheet.write(row, 11, realtor.address)
                    worksheet.write(row, 12, realtor.work_area)
                    worksheet.write(row, 13, realtor.status)
                    worksheet.write(row, 14, realtor.is_cooperate)
                    worksheet.write(row, 15, realtor.is_published)
                    worksheet.write(row, 16, realtor.training)
                    worksheet.write(row, 17, realtor.referral)

                    # incrementing the value of row by one
                    # with each iterations.
                    row += 1

                workbook.close()
                output.seek(0)
                resp = HttpResponse(output.read(), content_type=content_type)
                resp[
                    'Content-Disposition'] = f'filename=Realtor_export-{datetime.today().strftime("%Y-%m-%d")}.xlsx'
                output.close()
                return resp

        context = self.get_export_context_data()
        context.update(self.admin_site.each_context(request))

        context['title'] = _("Export")
        context['form'] = form
        context['opts'] = self.model._meta
        request.current_app = self.admin_site.name
        return TemplateResponse(request, [self.export_template_name],
                                context)

    make_published.short_description = _("Công khai danh tính")
    unpublished.short_description = _("Ẩn danh tính")

admin.site.register(Realtor, RealtorAdmin)
