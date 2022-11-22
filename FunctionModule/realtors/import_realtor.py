import logging
import xlsxwriter
from django.db.models import Q
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from FunctionModule.listings.helpers import print_trace
from FunctionModule.realtors.choices import Position, Title, Workplace, Status
from FunctionModule.realtors.models import Realtor

default_password = 'tgnpvn@2021'
logger = logging.getLogger(__name__)


def get_columns_from_worksheet(ws):
    return {
        cell.value: {
            'letter': get_column_letter(cell.column),
            'number': cell.column - 1
        } for cell in ws[1] if cell.value
    }


def read_header(header_row):
    header_dict = {
        "tieude": 0,
        "mota": 1,
        "mota-dauchu": 2,
        "anh-nha": 3,
        "anh-so": 4,
        "tgian": 5,
        "hien-trang": 6,
        "dia-chi": 7,
        "pho": 8,
        "quan": 9,
        "thong-so": 10,
        "gia": 11,
        "dau-chu": 12,
        "sdt": 13,
        "don-vi": 14,
        "dac-diem": 15,
        "huong": 16,
        "dt": 20,
        "trm2": 21,
        "hientrang": 22,
        # "thanh-pho": 20,
        # "so-tang": 23,
        # "mat-tien": 23,
        # "hoa-hong": 23,
        # "nguon": 23,
    }

    return header_dict


def get_direction(direction: str) -> str:
    code = ''
    if direction == 'Đông':
        code = 'east'
    elif direction == 'Tây':
        code = 'west'
    elif direction == 'Nam':
        code = 'south'
    elif direction == 'Bắc':
        code = 'north'
    elif direction == 'Đông Nam':
        code = 'south-east'
    elif direction == 'Đông Bắc':
        code = 'north-east'
    elif direction == 'Tây Nam':
        code = 'south-west'
    elif direction == 'Tây Bắc':
        code = 'north-west'

    return code


def handle_import(request, file_path):
    line_count = 0
    try:
        if request.user is not None:
            user = request.user
            query = Q(phone1=user.phone) or Q(user_id=user.id)
            real = Realtor.objects.filter(query)
            if real.first() is None:
                Realtor.objects.create(user_id=user.id, phone1=user.phone, name=user.name, email=user.email, address=user.address, is_cooperate=True)

        workbook = load_workbook(file_path)
        # Define variable to read the active sheet:
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            line_count += 1
            realtor = Realtor()
            cols = [c.value for c in row]
            realtor.name = cols[0]
            if cols[1] is not None:
                year = cols[1].strftime("%Y")
                realtor.birthyear = int(str(year))
            if cols[2] is not None:
                phone = cols[2].split(' ')
                realtor.phone1 = phone[0]
                if len(phone) > 1:
                    realtor.phone2 = phone[1]
            else:
                realtor.phone1 = realtor.name
            realtor.identifier = cols[3]
            if cols[4] is not None:
                realtor.date_join = cols[4].strftime("%Y-%m-%d")
            if cols[5] is not None:
                if "trợ lý" in cols[5].lower():
                    realtor.position = Position.ASSISTANT
                    realtor.title = Title.COMPETENTLY
                    realtor.level = 5
                elif "thư ký" in cols[5].lower():
                    realtor.position = Position.SECRETARY
                    realtor.title = Title.COMPETENTLY
                    realtor.level = 6
                elif "đầu chủ" in cols[5].lower() or "ứng viên đầu chủ" in cols[5].lower():
                    realtor.position = Position.EXPERT_HOME
                    realtor.title = Title.MASTER
                    realtor.level = 7
                elif "giám đốc" in cols[5].lower():
                    realtor.position = Position.MANAGING_DIRECTOR
                    realtor.title = Title.EXPERT
                    realtor.level = 9
                elif "trưởng phòng" in cols[5].lower() or "uvtp" in cols[5].lower() or "ứng viên trưởng phòng" in cols[
                    5].lower():
                    realtor.position = Position.MANAGER
                    realtor.title = Title.EXPERT
                    realtor.level = 8
            if cols[6] is not None:
                if "nhà phố việt nam" in cols[6].lower():
                    realtor.workplace = Workplace.NHAPHO
                elif "thế giới nhà phố" in cols[6].lower():
                    realtor.workplace = Workplace.TGNP
                elif "thiên khôi" in cols[6].lower():
                    realtor.workplace = Workplace.THIENKHOI
                elif "tuấn 123" in cols[6].lower():
                    realtor.workplace = Workplace.TUAN123
                elif "kimland" in cols[6].lower():
                    realtor.workplace = Workplace.KIMLAND
                elif "sumo" in cols[6].lower():
                    realtor.workplace = Workplace.SUMO
                elif "vuda" in cols[6].lower():
                    realtor.workplace = Workplace.VUDA
                elif "khác" in cols[6].lower():
                    realtor.workplace = Workplace.OTHER
            realtor.department = cols[7]
            realtor.facebook = cols[8]
            realtor.email = cols[9]
            realtor.countryside = cols[10]
            realtor.address = cols[11]
            realtor.work_area = cols[12]
            if cols[13] is not None:
                if "đang hợp tác" in cols[13].lower():
                    realtor.status = Status.COOPERATING
                elif "tạm dừng" in cols[13].lower():
                    realtor.status = Status.PAUSE_COOPERATING
                elif "ngừng hợp tác" in cols[13].lower() or "dừng hợp tác" in cols[13].lower():
                    realtor.status = Status.STOP_COOPERATING
            if cols[14] is not None:
                if "đang hợp tác" in cols[14].lower():
                    realtor.is_cooperate = True
                else:
                    realtor.is_cooperate = False
            if cols[15] is not None:
                if "có" in cols[15].lower():
                    realtor.is_published = True
                else:
                    realtor.is_cooperate = False
            realtor.training = cols[16]
            realtor.referral = cols[17]

            query = Q(name=realtor.name)
            if realtor.email is not None:
                query = query or Q(email=realtor.email)
            if realtor.facebook is not None:
                query = query or Q(facebook=realtor.facebook)

            if realtor.phone1 != realtor.name:
                if not Realtor.objects.filter(phone1=realtor.phone1).exists():
                    if realtor.phone1 is not None:
                        realtor.save()
                        print(f"Thêm Realtor {realtor} chưa có")
                else:
                    real = Realtor.objects.filter(phone1=realtor.phone1).first()
                    real.name = realtor.name
                    real.phone1 = realtor.phone1
                    real.identifier = realtor.identifier
                    real.birthyear = realtor.birthyear
                    real.date_join = realtor.date_join
                    real.position = realtor.position
                    real.workplace = realtor.workplace
                    real.department = realtor.department
                    real.facebook = realtor.facebook
                    real.email = realtor.email
                    real.countryside = realtor.countryside
                    real.work_area = realtor.work_area
                    real.status = realtor.status
                    real.is_cooperate = realtor.is_cooperate
                    real.is_published = realtor.is_published
                    real.save()
                    print(f"Cập nhật realtor {real}")
            elif not Realtor.objects.filter(query).exists() and realtor.phone1 is not None:
                realtor.save()
                print(f"Thêm Realtor {realtor} chưa có số")
            elif realtor.phone1 == realtor.name:
                reals = Realtor.objects.filter(facebook=realtor.facebook)
                if reals.count() > 1:
                    for r in reals:
                        if not r.phone1.isalnum():
                            r.delete()

        workbook.close()

    except Exception as ex:
        print(f"Error occurred at line: {line_count} type {type(ex)}")
        print_trace(ex)
    finally:
        print(f'Read {line_count} lines')
